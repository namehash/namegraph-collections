from argparse import ArgumentParser
from pathlib import Path
from itertools import count
import json

import openpyxl
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.cell.read_only import ReadOnlyCell


def collect_kwargs(worksheet: ReadOnlyWorksheet) -> dict:
    kwargs = dict()
    collecting_labels = False

    for row_idx in count(1):
        key_cell: ReadOnlyCell = worksheet.cell(row_idx, 1)
        value_cell: ReadOnlyCell = worksheet.cell(row_idx, 2)

        if key_cell.value == 'normalized_label' and value_cell.value == 'tokenized_label':
            kwargs['labels'] = []
            collecting_labels = True
            continue

        if not collecting_labels:
            if key_cell.value == 'collection_keywords':
                keywords = []
                for col_idx in count(2):
                    cell: ReadOnlyCell = worksheet.cell(row_idx, col_idx)
                    cell_value = cell.value.strip() if cell.value is not None else None
                    if not cell_value:
                        break
                    keywords.append(cell_value)
                kwargs['collection_keywords'] = keywords
                continue

            cell_value = value_cell.value.strip() if value_cell.value is not None else None
            kwargs[key_cell.value] = cell_value

        if collecting_labels:
            normalized_label = key_cell.value.strip() if key_cell.value is not None else None
            tokenized_label = value_cell.value.strip() if value_cell.value is not None else None

            if not normalized_label and not tokenized_label:
                break

            if tokenized_label and not normalized_label:
                normalized_label = ''.join(tokenized_label.split(' '))

            label_dictionary = {
                'normalized_label': normalized_label
            }

            if tokenized_label:
                label_dictionary['tokenized_label'] = tokenized_label.split(' ')

            kwargs['labels'].append(label_dictionary)

    return kwargs


def process_sheet(worksheet: ReadOnlyWorksheet) -> dict:
    kwargs = collect_kwargs(worksheet)
    collection = {
        'commands': {},
        'data': {
            'collection_id': kwargs['collection_id'],
            'collection_name': kwargs['collection_name'],
            'labels': kwargs['labels']
        }
    }

    if kwargs['sort_labels']:
        collection['commands']['sort_labels'] = kwargs['sort_labels']
    if kwargs['collection_description']:
        collection['data']['collection_description'] = kwargs['collection_description']
    if kwargs['collection_keywords']:
        collection['data']['collection_keywords'] = kwargs['collection_keywords']
    if kwargs['avatar_emoji']:
        collection['data']['avatar_emoji'] = kwargs['avatar_emoji']

    return collection


if __name__ == '__main__':
    parser = ArgumentParser()
    parser.add_argument('--input', type=str, help='input XLSX file with custom collections')
    parser.add_argument('--output-dir', type=str, help='output directory for the JSON files')
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    wb: openpyxl.Workbook = openpyxl.load_workbook(args.input, read_only=True)

    for worksheet in wb.worksheets:
        if worksheet.title == 'Template':
            continue

        collection = process_sheet(worksheet)
        with open(output_dir / (worksheet.title + '.json'), 'w') as f:
            json.dump(collection, f, indent=2, ensure_ascii=False)
